#!/usr/bin/env python 
# -*- coding: utf-8 -*- 
# @Time : 2019/3/25 11:39 
# @Author : wangchong 
# @Email: chongwangcc@gmail.com
# @File : ExcelTools.py 
# @Software: PyCharm
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment,Border, Side


class BalanceSheetTools:
    """
    写财报的工具
    """
    bold_font = Font(bold=True)
    left, right, top, bottom = [Side(style='medium', color='000000')] * 4
    border = Border(left=left, right=right, top=top, bottom=bottom)
    up_border = Border(top=top)
    left_border = Border(left=left)
    right_border = Border(right=right)
    bottom_border = Border(bottom=bottom)
    left_top_border = Border(top=top, left=left)
    right_top_border = Border(top=top, right=right)
    left_bottom_border = Border(bottom=bottom, left=left)
    right_bottom_border = Border(bottom=bottom, right=right)

    center_alignment = Alignment(horizontal='center', vertical='center')

    def __init__(self, excel_path):
        """
        初始化
        :param excel_path:
        """
        self.excel_path = excel_path
        self.sheet_row_start = {} # 数据写到第几行了
        self.wb = Workbook()

    def __get_start_cell(self, sheet_name):
        """
        获得sheet下一个可写入数据的cell
        :param sheet_name:
        :return:
        """
        default_row = self.sheet_row_start.setdefault(sheet_name, 2)
        if sheet_name not in self.wb.sheetnames:
            ws = self.wb.create_sheet(sheet_name)
        else:
            ws = self.wb[sheet_name]
        return ws, default_row

    def append_report_sheet(self, sheet_name, report_data):
        """
        添加一个报表
        :param sheet_name:
        :param report_data:
        :return:
        """
        # 1. 合并单元格，设置表头
        ws, cell_start = self.__get_start_cell(sheet_name)
        ws.merge_cells(start_row=cell_start, start_column=1, end_row=cell_start, end_column=6)
        ws.cell(row=cell_start, column=1, value=report_data["title_center"]).font = self.bold_font
        ws.cell(row=cell_start, column=1).alignment = self.center_alignment
        ws.merge_cells(start_row=cell_start+1, start_column=1, end_row=cell_start+1, end_column=3)
        ws.cell(row=cell_start+1, column=1, value=report_data["title_left"]).font = self.bold_font
        ws.cell(row=cell_start+1, column=1).alignment = self.center_alignment
        ws.merge_cells(start_row=cell_start + 1, start_column=4, end_row=cell_start + 1, end_column=6)
        ws.cell(row=cell_start + 1, column=4, value=report_data["title_right"]).font = self.bold_font
        ws.cell(row=cell_start + 1, column=4).alignment = self.center_alignment
        ws.cell(row=cell_start + 2, column=1, value="名称")
        ws.cell(row=cell_start + 2, column=2, value="原价（元）")
        ws.cell(row=cell_start + 2, column=3, value="现价（元）")
        ws.cell(row=cell_start + 2, column=4, value="名称")
        ws.cell(row=cell_start + 2, column=5, value="原价（元）")
        ws.cell(row=cell_start + 2, column=6, value="现价（元）")

        # 2. 设置 左半边 表格
        left_row_nums = cell_start + 3
        for t_data in report_data["data_left"]["data"]:
            for c in range(1, 4):
                ws.cell(row=left_row_nums, column=c).border = self.up_border
            ws.cell(row=left_row_nums, column=1, value=t_data["name"]).font = self.bold_font
            ws.cell(row=left_row_nums, column=2, value=t_data["sum_org"]).font = self.bold_font
            ws.cell(row=left_row_nums, column=3, value=t_data["sum_now"]).font = self.bold_font
            ws.cell(row=left_row_nums, column=1).border = self.left_top_border
            ws.cell(row=left_row_nums, column=3).border = self.right_top_border
            left_row_nums += 1
            if "data" in t_data.keys():
                for tt_data in t_data["data"]:
                    ws.cell(row=left_row_nums, column=1, value=("    "+tt_data["name"]))
                    ws.cell(row=left_row_nums, column=2, value=tt_data["sum_org"])
                    ws.cell(row=left_row_nums, column=3, value=tt_data["sum_now"])
                    ws.cell(row=left_row_nums, column=1).border = self.left_border
                    ws.cell(row=left_row_nums, column=3).border = self.right_border
                    left_row_nums += 1
            ws.cell(row=left_row_nums, column=1, value="")
            ws.cell(row=left_row_nums, column=2, value="")
            ws.cell(row=left_row_nums, column=3, value="")
            ws.cell(row=left_row_nums, column=1).border = self.left_border
            ws.cell(row=left_row_nums, column=3).border = self.right_border
            left_row_nums += 1

        # 3. 设置 右半边 表格
        right_row_nums = cell_start + 3
        for t_data in report_data["data_right"]["data"]:
            for c in range(4, 7):
                ws.cell(row=right_row_nums, column=c).border = self.up_border
            ws.cell(row=right_row_nums, column=4, value=t_data["name"]).font = self.bold_font
            ws.cell(row=right_row_nums, column=5, value=t_data["sum_org"]).font = self.bold_font
            ws.cell(row=right_row_nums, column=6, value=t_data["sum_now"]).font = self.bold_font
            ws.cell(row=right_row_nums, column=4).border = self.left_top_border
            ws.cell(row=right_row_nums, column=6).border = self.right_top_border
            right_row_nums += 1
            if "data" in t_data.keys():
                for tt_data in t_data["data"]:
                    ws.cell(row=right_row_nums, column=4, value=("    "+tt_data["name"]))
                    ws.cell(row=right_row_nums, column=5, value=tt_data["sum_org"])
                    ws.cell(row=right_row_nums, column=6, value=tt_data["sum_now"])
                    ws.cell(row=right_row_nums, column=4).border = self.left_border
                    ws.cell(row=right_row_nums, column=6).border = self.right_border
                    right_row_nums += 1
            ws.cell(row=right_row_nums, column=4, value="")
            ws.cell(row=right_row_nums, column=5, value="")
            ws.cell(row=right_row_nums, column=6, value="")
            ws.cell(row=right_row_nums, column=4).border = self.left_border
            ws.cell(row=right_row_nums, column=6).border = self.right_border
            right_row_nums += 1

        # 4. 设置 底部 汇总信息
        all_nums = left_row_nums if left_row_nums > right_row_nums else right_row_nums
        ws.cell(row=all_nums, column=1, value=report_data["data_left"]["name"]).font = self.bold_font
        ws.cell(row=all_nums, column=2, value=report_data["data_left"]["sum_org"]).font = self.bold_font
        ws.cell(row=all_nums, column=3, value=report_data["data_left"]["sum_now"]).font = self.bold_font
        ws.cell(row=all_nums, column=4, value=report_data["data_right"]["name"]).font = self.bold_font
        ws.cell(row=all_nums, column=5, value=report_data["data_right"]["sum_org"]).font = self.bold_font
        ws.cell(row=all_nums, column=6, value=report_data["data_right"]["sum_now"]).font = self.bold_font

        # 5. 设置excel 样式 边框
        for c in range(1, 7):
            ws.cell(row=cell_start, column=c).border = self.border
            ws.cell(row=cell_start + 1, column=c).border = self.border
            ws.cell(row=cell_start + 2, column=c).border = self.border
            ws.cell(row=all_nums, column=c).border = self.border

        for c in "ABCDEF":
            ws.column_dimensions[c].width = 18

        # 6. 更新 最后写入行
        self.sheet_row_start[sheet_name] = all_nums+3

    def append_table(self, sheet_name, table_data):
        """
        追加 附表
        :param sheet_name:
        :param table_data:
        :return:
        """
        ws, cell_start = self.__get_start_cell(sheet_name)
        columns_nums = len(table_data["columns"])

        # 1.创建表头
        ws.merge_cells(start_row=cell_start, start_column=1, end_row=cell_start, end_column=columns_nums)
        ws.cell(row=cell_start, column=1, value=table_data["title"])
        ws.cell(row=cell_start, column=1).font = self.bold_font
        ws.cell(row=cell_start, column=1).alignment = self.center_alignment
        for c in range(1,columns_nums+1):
            ws.cell(row=cell_start, column=c).border = self.border
        cell_start += 1

        for index, c in enumerate(table_data["columns"]):
            ws.cell(row=cell_start , column=index + 1, value=c)
            ws.cell(row=cell_start, column=index + 1).font = self.bold_font
            ws.cell(row=cell_start, column=index + 1).alignment = self.center_alignment
            ws.cell(row=cell_start, column=index + 1).border = self.border
        cell_start += 1
        # 2. 设置数据
        for index1, t_data in enumerate(table_data["data"]):
            for index, c in enumerate(table_data["columns"]):
                ws.cell(row=cell_start, column=index + 1, value=t_data[c])
                if index == 0:
                    ws.cell(row=cell_start, column=index + 1).border = self.left_border
                    pass
                elif index == (columns_nums -1):
                    ws.cell(row=cell_start, column=index + 1).border = self.right_border
                    pass
            cell_start += 1
        for index in range(0, columns_nums):
                ws.cell(row=cell_start, column=index + 1).border = self.up_border

        # 6. 更新 最后写入行
        self.sheet_row_start[sheet_name] = cell_start + 2

    def save(self):
        self.wb.save(self.excel_path)


if __name__ == "__main__":
    report_data = {
        "title_center": "损益表",
        "title_left": "收入（元）",
        "title_right": "支出（元）",
        "data_left": {
            "name": "总收入",
            "sum_org": 12398.98,
            "sum_now": 1290.89,
            "data": [
                {
                    "name": "E象限",
                    "sum_org": 12398.98,
                    "sum_now": 1290.98,
                    "data": [
                        {
                            "name": "工资(CC)",
                            "sum_org": 12830.0,
                            "sum_now": 12908.0,
                        },{
                            "name": "工资(MM)",
                            "sum_org": 12830.0,
                            "sum_now": 12908.0,
                        }
                    ]
                },{
                    "name": "S象限",
                    "sum_org": 12398.98,
                    "sum_now": 1290.98,
                    "data": [
                        {
                            "name": "比赛",
                            "sum_org": 30.0,
                            "sum_now": 18.0,
                        }, {
                            "name": "捡钱",
                            "sum_org": 2830.0,
                            "sum_now": 2908.0,
                        }
                    ]
                },{
                    "name": "B象限",
                    "sum_org": 0,
                    "sum_now": 0,
                },{
                    "name": "I象限",
                    "sum_org": 12398.98,
                    "sum_now": 1290.98,
                    "data": [
                        {
                            "name": "利息",
                            "sum_org": 30.0,
                            "sum_now": 18.0,
                        }, {
                            "name": "股票",
                            "sum_org": 2830.0,
                            "sum_now": 2908.0,
                        }
                    ]
                }
            ]
        },
        "data_right": {
            "name": "总支出",
            "sum_org": 12398.98,
            "sum_now": 1290.89,
            "data": [
                {
                    "name": "消费",
                    "sum_org": 12398.98,
                    "sum_now": 1290.98,
                    "data": [
                        {
                            "name": "生活费账户",
                            "sum_org": 12830.0,
                            "sum_now": 12908.0,
                        }, {
                            "name": "doodads账户",
                            "sum_org": 12830.0,
                            "sum_now": 12908.0,
                        }, {
                            "name": "住房账户",
                            "sum_org": 12830.0,
                            "sum_now": 12908.0,
                        }, {
                            "name": "教育基金",
                            "sum_org": 12830.0,
                            "sum_now": 12908.0,
                        }, {
                            "name": "风险备付金",
                            "sum_org": 12830.0,
                            "sum_now": 12908.0,
                        }
                    ]
                }, {
                    "name": "投资",
                    "sum_org": 12398.98,
                    "sum_now": 1290.98,
                    "data": [
                        {
                            "name": "股票账户",
                            "sum_org": 30.0,
                            "sum_now": 18.0,
                        }, {
                            "name": "基金定投",
                            "sum_org": 2830.0,
                            "sum_now": 2908.0,
                        }, {
                            "name": "杠杆利息",
                            "sum_org": 2830.0,
                            "sum_now": 2908.0,
                        }
                    ]
                }, {
                    "name": "储蓄账户",
                    "sum_org": 0,
                    "sum_now": 0,
                }
            ]
        }
    }
    table_data = {
        "title": "银行卡明细表",
        "columns": ["名称", "原价", "现价", "备注", "所属账户"],
        "data": [
            {"名称": "建行卡（CC）", "原价": 12367, "现价": 789, "备注": "打发", "所属账户": "花销账户-生活费"},
            {"名称": "建行卡（CC）", "原价": 12367, "现价": 789, "备注": "打发", "所属账户": "花销账户-生活费"},
            {"名称": "建行卡（CC）", "原价": 12367, "现价": 789, "备注": "打发", "所属账户": "花销账户-生活费"},
            {"名称": "建行卡（CC）", "原价": 12367, "现价": 789, "备注": "打发", "所属账户": "花销账户-生活费"},
            {"名称": "建行卡（CC）", "原价": 12367, "现价": 789, "备注": "打发", "所属账户": "花销账户-生活费"},
        ]

    }

    sheet_name = "合并财报"
    tools = BalanceSheetTools(r"./data/reports/test.xlsx")
    tools.append_report_sheet(sheet_name, report_data)
    tools.append_report_sheet(sheet_name, report_data)
    tools.append_table(sheet_name, table_data)
    tools.append_table(sheet_name, table_data)
    tools.save()






